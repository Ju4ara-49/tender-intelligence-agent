"""Интеграционный тест Excel-экспортера на реальной SQLite-схеме проекта."""

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from src.export.excel import export_tenders_to_excel
from src.models.tender import Tender, TenderAnalysis
from src.storage.database import TenderDatabase


def main() -> None:
    output = Path("output/github_actions_excel_test.xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)

    db_path = Path("output/github_actions_test.db")
    if db_path.exists():
        db_path.unlink()

    db = TenderDatabase(db_path)

    tender = Tender(
        platform="b2b_center",
        external_id="GHA-TEST-001",
        title="Поставка подшипников — интеграционный тест",
        url="https://example.com/tender/GHA-TEST-001",
        description="Тестовая закупка для проверки полного Excel-экспортера.",
        price=125000.0,
        currency="RUB",
        deadline=datetime(2026, 9, 1, 12, 0, tzinfo=timezone.utc),
        published_at=datetime(2026, 8, 22, 10, 0, tzinfo=timezone.utc),
        region="Москва",
        customer="Тестовый заказчик",
        law_type="223-ФЗ",
        raw_data={"procurement_method": "Запрос предложений"},
    )

    tender_id = db.save_tender(tender)

    db.save_analysis(
        tender_id,
        TenderAnalysis(
            relevance_score=85,
            summary="Тестовое резюме: тендер подходит для проверки экспорта.",
            recommendation="review",
            risks=["Тестовый риск"],
            deadline_note="Осталось достаточно времени для подготовки заявки.",
            is_stub=False,
        ),
    )

    export_tenders_to_excel(db, output, tender_ids=[tender_id])

    wb = load_workbook(output)
    ws = wb["Тендеры"]

    headers = [cell.value for cell in ws[1]]
    expected = [
        "Площадка", "Номер закупки", "Наименование", "Заказчик", "Регион",
        "НМЦК", "Валюта", "Дата публикации", "Дата окончания подачи заявок",
        "Осталось дней до подачи", "Закон", "Способ закупки", "AI score",
        "Рекомендация", "Краткое резюме", "Риски", "Начальная цена",
        "Комментарий по срокам", "Ссылка",
    ]

    assert headers == expected, headers
    assert "Дата поиска" not in headers
    assert "№ поиска" not in headers
    assert "Статус" not in headers
    assert "W" not in headers
    assert "X" not in headers
    assert "Y" not in headers

    assert ws["Q2"].value == 125000.0
    assert ws["L2"].value == "Запрос предложений"
    assert ws["R2"].value == "Осталось достаточно времени для подготовки заявки."

    hyperlink = ws["S2"].hyperlink
    assert hyperlink is not None
    assert hyperlink.target == "https://example.com/tender/GHA-TEST-001"

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == ws.dimensions

    print("INTEGRATION TEST: PASS")
    print(f"Excel: {output}")
    print(f"Rows: {ws.max_row - 1}")
    print(f"Columns: {ws.max_column}")
    print(f"Clickable tender link: {hyperlink.target}")
    print(f"Initial price: {ws['Q2'].value}")
    print(f"Procurement method: {ws['L2'].value}")
    print("Excel schema checks: PASS")


if __name__ == "__main__":
    main()
