"""Многопользовательский runtime Telegram-бота с белым списком доступа."""
from __future__ import annotations

import json
import logging
import os
import threading
import time
from html import escape
from pathlib import Path

from openpyxl import load_workbook

from src.orchestrator import Orchestrator
from src.telegram_bot import TelegramBot

logger = logging.getLogger(__name__)

OWNER_TELEGRAM_ID = "838120236"
WHITELIST_FILE = Path("data/telegram_allowed_users.json")

BTN_ADMIN = "👑 Управление доступом"
BTN_ADMIN_ADD = "➕ Добавить пользователя"
BTN_ADMIN_REMOVE = "➖ Удалить пользователя"
BTN_ADMIN_USERS = "📋 Список пользователей"
BTN_ADMIN_BACK = "↩️ Назад"


class MultiUserTelegramBot(TelegramBot):
    """TelegramBot с независимыми поисками и управлением белым списком."""

    def __init__(self, settings, orchestrator: Orchestrator) -> None:
        super().__init__(settings, orchestrator)
        self._search_threads: dict[str, threading.Thread] = {}
        self._user_orchestrators: dict[str, Orchestrator] = {}
        self._search_lock = threading.Lock()
        self._whitelist_lock = threading.Lock()
        self._admin_waiting: dict[str, str] = {}
        self._allowed_user_ids = self._load_allowed_user_ids()
        logger.info(
            "Telegram-доступ: whitelist включён; разрешённых пользователей=%d",
            len(self._allowed_user_ids),
        )

    def _load_allowed_user_ids(self) -> set[str]:
        allowed = {
            item.strip()
            for item in os.getenv("TELEGRAM_ALLOWED_USER_IDS", "").split(",")
            if item.strip()
        }
        try:
            if WHITELIST_FILE.exists():
                data = json.loads(WHITELIST_FILE.read_text(encoding="utf-8"))
                if isinstance(data, list):
                    allowed.update(str(x).strip() for x in data if str(x).strip())
        except Exception:
            logger.exception("Telegram-доступ: не удалось прочитать %s", WHITELIST_FILE)
        allowed.add(OWNER_TELEGRAM_ID)
        if self.admin_chat_id:
            allowed.add(self.admin_chat_id)
        return allowed

    def _save_allowed_user_ids(self) -> None:
        WHITELIST_FILE.parent.mkdir(parents=True, exist_ok=True)
        users = sorted(x for x in self._allowed_user_ids if x != OWNER_TELEGRAM_ID)
        WHITELIST_FILE.write_text(
            json.dumps(users, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _is_owner(self, chat_id: str) -> bool:
        return str(chat_id).strip() == OWNER_TELEGRAM_ID

    def _is_allowed(self, chat_id: str) -> bool:
        return str(chat_id).strip() in self._allowed_user_ids

    def _access_denied(self, chat_id: str) -> None:
        user_id = str(chat_id).strip()
        self._send(
            chat_id,
            "🔒 <b>Доступ ограничен.</b>\n\n"
            "У вас пока нет разрешения на использование этого бота.\n\n"
            f"Ваш Telegram ID: <code>{user_id}</code>\n\n"
            "Передайте этот ID администратору. После добавления в белый список "
            "бот станет доступен вам автоматически.",
        )

    def _admin_keyboard(self) -> dict:
        return {
            "keyboard": [
                [{"text": BTN_ADMIN_ADD}],
                [{"text": BTN_ADMIN_REMOVE}],
                [{"text": BTN_ADMIN_USERS}],
                [{"text": BTN_ADMIN_BACK}],
            ],
            "resize_keyboard": True,
            "is_persistent": True,
        }

    def _admin_menu(self, chat_id: str) -> None:
        self._send(
            chat_id,
            "<b>👑 Управление доступом</b>\n\n"
            "Выберите действие:",
            self._admin_keyboard(),
        )

    def _show_users(self, chat_id: str) -> None:
        users = sorted(self._allowed_user_ids)
        lines = ["<b>👥 Разрешённые пользователи</b>", "", f"Всего: {len(users)}", ""]
        for user_id in users:
            suffix = " — владелец" if user_id == OWNER_TELEGRAM_ID else ""
            lines.append(f"• <code>{user_id}</code>{suffix}")
        self._send(chat_id, "\n".join(lines), self._admin_keyboard())

    def _admin_command(self, chat_id: str, text: str) -> bool:
        if not self._is_owner(chat_id):
            return False
        text = text.strip()
        parts = text.split()
        command = parts[0].lower() if parts else ""

        if command == "/admin":
            self._admin_waiting.pop(chat_id, None)
            self._admin_menu(chat_id)
            return True
        if command in {"/users", "/пользователи"}:
            self._admin_waiting.pop(chat_id, None)
            self._show_users(chat_id)
            return True
        if command in {"/add_user", "/добавить"}:
            if len(parts) != 2 or not parts[1].isdigit():
                self._admin_waiting[chat_id] = "add"
                self._send(chat_id, "Введите Telegram ID пользователя, которого нужно добавить:", self._admin_keyboard())
                return True
            self._add_user(chat_id, parts[1])
            return True
        if command in {"/remove_user", "/удалить"}:
            if len(parts) != 2 or not parts[1].isdigit():
                self._admin_waiting[chat_id] = "remove"
                self._send(chat_id, "Введите Telegram ID пользователя, которого нужно удалить:", self._admin_keyboard())
                return True
            self._remove_user(chat_id, parts[1])
            return True
        return False

    def _add_user(self, chat_id: str, user_id: str) -> None:
        with self._whitelist_lock:
            self._allowed_user_ids.add(user_id)
            self._save_allowed_user_ids()
        self._admin_waiting.pop(chat_id, None)
        self._send(
            chat_id,
            f"✅ Пользователь <code>{user_id}</code> добавлен в белый список.",
            self._admin_keyboard(),
        )

    def _remove_user(self, chat_id: str, user_id: str) -> None:
        if user_id == OWNER_TELEGRAM_ID:
            self._send(chat_id, "⛔ Владельца удалить нельзя. Ваш доступ постоянный.", self._admin_keyboard())
            return
        with self._whitelist_lock:
            self._allowed_user_ids.discard(user_id)
            self._save_allowed_user_ids()
        self._admin_waiting.pop(chat_id, None)
        self._send(
            chat_id,
            f"✅ Пользователь <code>{user_id}</code> удалён из белого списка.",
            self._admin_keyboard(),
        )

    def _handle_message(self, message: dict) -> None:
        chat_id = str(message.get("chat", {}).get("id", ""))
        text = (message.get("text") or "").strip()
        if not chat_id:
            return

        if self._is_owner(chat_id) and chat_id in self._admin_waiting and text:
            if text == BTN_ADMIN_BACK:
                self._admin_waiting.pop(chat_id, None)
                self._send(chat_id, "Возвращаемся в основное меню.", self._keyboard())
                return
            if text in {BTN_ADMIN_ADD, BTN_ADMIN_REMOVE, BTN_ADMIN_USERS, BTN_ADMIN}:
                pass
            elif text.isdigit():
                action = self._admin_waiting.get(chat_id)
                if action == "add":
                    self._add_user(chat_id, text)
                else:
                    self._remove_user(chat_id, text)
                return
            else:
                self._send(chat_id, "Нужен числовой Telegram ID. Например: <code>1378791558</code>", self._admin_keyboard())
                return

        if self._is_owner(chat_id):
            if text == BTN_ADMIN:
                self._admin_menu(chat_id)
                return
            if text == BTN_ADMIN_ADD:
                self._admin_waiting[chat_id] = "add"
                self._send(chat_id, "Введите Telegram ID пользователя, которого нужно добавить:", self._admin_keyboard())
                return
            if text == BTN_ADMIN_REMOVE:
                self._admin_waiting[chat_id] = "remove"
                self._send(chat_id, "Введите Telegram ID пользователя, которого нужно удалить:", self._admin_keyboard())
                return
            if text == BTN_ADMIN_USERS:
                self._show_users(chat_id)
                return
            if text == BTN_ADMIN_BACK:
                self._send(chat_id, "Возвращаемся в основное меню.", self._keyboard())
                return
            if self._admin_command(chat_id, text):
                return

        if not self._is_allowed(chat_id):
            logger.warning("Telegram-доступ: отказ пользователю chat_id=%s", chat_id)
            self._access_denied(chat_id)
            return

        # Кнопка «Ключевые слова» теперь сразу переводит пользователя
        # в режим ввода. Дополнительно свободный текст из основного меню
        # принимаем как ключевые слова — это восстанавливает привычный
        # сценарий: можно просто написать «станки».
        if text == self.BTN_KEYWORDS if hasattr(self, "BTN_KEYWORDS") else False:
            pass

        if text == "Ключевые слова":
            self._ask_value(
                chat_id,
                "keywords",
                "<b>Введите ключевые слова</b> через запятую.\n\n"
                "Например:\n<code>станки, токарные станки, фрезерные станки</code>\n\n"
                "Можно ввести одно слово, например: <code>станки</code>.",
            )
            return

        menu_buttons = {
            "Цена от:", "Цена до:", "Балл:", "Срок:", "Ключевые слова",
            "Площадки", "Сброс", "Поиск", "Настройки", "Стоп", "Помощь",
        }

        if text and not text.startswith("/") and text not in menu_buttons and chat_id not in self._waiting_for:
            self._ask_value(chat_id, "keywords", "Ключевые слова сохранены.\n\nВведите ключевые слова через запятую:")
            self._handle_value_input(chat_id, text)
            return

        super()._handle_message(message)

    def _handle_callback(self, callback: dict) -> None:
        message = callback.get("message") or {}
        chat_id = str(message.get("chat", {}).get("id", ""))
        if chat_id and not self._is_allowed(chat_id):
            self._answer_callback(str(callback.get("id", "")))
            self._access_denied(chat_id)
            return
        super()._handle_callback(callback)

    def _cmd_search(self, chat_id: str) -> None:
        with self._search_lock:
            thread = self._search_threads.get(chat_id)
            if thread is not None and thread.is_alive():
                self._send(chat_id, "Поиск уже выполняется для вашего аккаунта.\n\nЕсли нужно остановить его — нажмите «Стоп».", self._keyboard())
                return
            orchestrator = Orchestrator(self.settings)
            orchestrator.notifier.chat_id = chat_id
            self._user_orchestrators[chat_id] = orchestrator
            orchestrator.clear_stop_request()
            self._send(chat_id, "Запускаю новый поиск тендеров.\n\nПоиск выполняется в фоне.", self._keyboard())
            thread = threading.Thread(target=self._run_search_for_user, args=(chat_id, orchestrator), daemon=True, name=f"telegram-search-{chat_id}")
            self._search_threads[chat_id] = thread
            thread.start()

    def _find_excel_results(self, search_number: int) -> list[dict[str, str]]:
        output_dir = Path(self.settings.config.get("export", {}).get("output_dir", "output"))
        files = sorted(output_dir.glob(f"search_{search_number:03d}_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not files:
            return []
        try:
            wb = load_workbook(files[0], read_only=True, data_only=True)
            ws = wb["Тендеры"]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                wb.close()
                return []
            headers = [str(x).strip() if x is not None else "" for x in rows[0]]
            index = {name: i for i, name in enumerate(headers)}
            results: list[dict[str, str]] = []
            for row in rows[1:]:
                title = str(row[index.get("Наименование", 2)] or "").strip()
                if not title:
                    continue
                results.append({
                    "title": title,
                    "price": str(row[index.get("Начальная цена", 5)] or "не указана"),
                    "deadline": str(row[index.get("Дата окончания подачи заявок", 8)] or "не указан"),
                    "score": str(row[index.get("AI score", 12)] or "—"),
                    "url": str(row[index.get("Ссылка", 16)] or ""),
                })
            wb.close()
            return results
        except Exception:
            logger.exception("Telegram: не удалось прочитать Excel результатов поиска №%03d", search_number)
            return []

    def _send_search_results(self, chat_id: str, results: list[dict[str, str]]) -> None:
        if not results:
            return
        chunks: list[str] = []
        current = "<b>📋 Результаты текущего поиска</b>\n\n"
        for number, item in enumerate(results, 1):
            title = escape(item["title"])
            price = escape(item["price"])
            deadline = escape(item["deadline"])
            score = escape(item["score"])
            url = item["url"].strip()
            link = f'\n🔗 <a href="{escape(url, quote=True)}">Открыть тендер</a>' if url else ""
            block = f"<b>{number}. {title}</b>\n💰 {price}\n⏰ {deadline} | AI: {score}{link}\n\n"
            if len(current) + len(block) > 3800:
                chunks.append(current.rstrip())
                current = "<b>📋 Продолжение результатов</b>\n\n"
            current += block
        if current.strip():
            chunks.append(current.rstrip())
        for chunk in chunks:
            self._send(chat_id, chunk, self._keyboard())

    def _run_search_for_user(self, chat_id: str, orchestrator: Orchestrator) -> None:
        started_at = time.monotonic()
        self._send(chat_id, "🔄 <b>Поиск выполняется...</b>\n\nИдёт сбор и анализ тендеров.\n\n⏳ Пожалуйста, подождите...", self._keyboard())
        try:
            stats = orchestrator.run_cycle()
            elapsed = int(time.monotonic() - started_at)
            elapsed_text = f"{elapsed // 60} мин. {elapsed % 60:02d} сек." if elapsed >= 60 else f"{elapsed} сек."
            state = "остановлен" if orchestrator.stop_requested else "завершён"
            results = self._find_excel_results(stats["search_number"])
            result_count = len(results)
            duplicate_note = " (все они уже встречались ранее)" if stats["skipped_duplicate"] and stats["new"] == 0 else ""
            text = (
                f"{'⛔' if orchestrator.stop_requested else '✅'} <b>Поиск №{stats['search_number']:03d} {state}.</b>\n\n"
                f"Время работы: {elapsed_text}\n\n"
                f"Найдено на площадках: {stats['found']}\n"
                f"Прошло фильтр по ключевым словам: {stats['filtered']}\n"
                f"Результатов в Excel: {result_count}\n"
                f"Новых тендеров: {stats['new']}\n"
                f"Проанализировано AI: {stats['analyzed']}\n"
                f"Исключено по критериям: {stats['excluded_by_criteria']}\n"
                f"Отправлено новых уведомлений: {stats['notified']}\n"
                f"Пропущено дублей: {stats['skipped_duplicate']}{duplicate_note}\n\n"
                "📊 <b>Результаты сохранены в Excel и показаны ниже.</b>"
            )
            self._send(chat_id, text, self._keyboard())
            self._send_search_results(chat_id, results)
        except Exception:
            logger.exception("Telegram-бот: ошибка выполнения поиска chat_id=%s", chat_id)
            self._send(chat_id, "❌ <b>Ошибка поиска.</b>\n\nПодробности находятся в logs/agent.log.", self._keyboard())
        finally:
            with self._search_lock:
                self._search_threads.pop(chat_id, None)
                self._user_orchestrators.pop(chat_id, None)

    def _cmd_stop(self, chat_id: str) -> None:
        with self._search_lock:
            orchestrator = self._user_orchestrators.get(chat_id)
            thread = self._search_threads.get(chat_id)
        if orchestrator is None or thread is None or not thread.is_alive():
            self._send(chat_id, "Сейчас ваш поиск не выполняется.", self._keyboard())
            return
        orchestrator.request_stop()
        self._send(chat_id, "Получена команда остановки. Текущий этап завершится, после чего ваш поиск будет остановлен.", self._keyboard())