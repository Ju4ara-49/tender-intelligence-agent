"""Экспорт результатов тендерного поиска в Excel."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_tenders_to_excel(
    db,
    output_path: Path | str,
    tender_ids: list[int] | None = None,
    search_number: int | None = None,
) -> Path:
    """Экспортирует результаты текущего прогона в отдельный Excel."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with db._connect() as conn:
        if tender_ids:
            placeholders = ",".join("?" for _ in tender_ids)

            rows = conn.execute(
                f"""
                SELECT
                    t.id,
                    t.platform,
                    t.external_id,
                    t.title,
                    t.url,
                    t.description,
                    t.price,
                    t.currency,
                    t.deadline,
                    t.published_at,
                    t.region,
                    t.customer,
                    t.law_type,
                    t.raw_data,

                    a.relevance_score,
                    a.summary,
                    a.recommendation,
                    a.risks,
                    a.deadline_note,
                    a.is_stub,
                    a.analyzed_at

                FROM tenders t
                LEFT JOIN analyses a
                    ON a.tender_id = t.id

                WHERE t.id IN ({placeholders})

                ORDER BY
                    CASE
                        WHEN a.relevance_score IS NULL THEN 0
                        ELSE a.relevance_score
                    END DESC,
                    t.published_at DESC
                """,
                tender_ids,
            ).fetchall()
        else:
            rows = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Тендеры"

    # Не выводим служебные поля поиска и статус.
    # W/X/Y из старой версии Excel также не формируются.
    headers = [
        "Площадка",
        "Номер закупки",
        "Наименование",
        "Заказчик",
        "Регион",
        "НМЦК",
        "Валюта",
        "Дата публикации",
        "Дата окончания подачи заявок",
        "Осталось дней до подачи",
        "Закон",
        "Способ закупки",
        "AI score",
        "Рекомендация",
        "Краткое резюме",
        "Риски",
        "Начальная цена",
        "Комментарий по срокам",
        "Ссылка",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    now = datetime.now(timezone.utc)

    for row in rows:
        raw_data = {}

        try:
            if row["raw_data"]:
                raw_data = json.loads(row["raw_data"])
        except (TypeError, ValueError, json.JSONDecodeError):
            raw_data = {}

        procurement_method = raw_data.get(
            "procurement_method",
            "",
        )

        platform = row["platform"] or ""

        platform_names = {
            "eis": "ЕИС",
            "b2b_center": "B2B-Center",
            "rts_tender": "РТС-тендер",
            "unipro": "Unipro",
            "tmk": "ТМК",
        }

        platform = platform_names.get(
            platform,
            platform,
        )

        recommendation_names = {
            "participate": "Участвовать",
            "review": "Рассмотреть",
            "skip": "Не участвовать",
        }

        recommendation = recommendation_names.get(
            row["recommendation"] or "",
            row["recommendation"] or "",
        )

        deadline = _parse_datetime(row["deadline"])
        published_at = _parse_datetime(row["published_at"])

        days_left = ""

        if deadline is not None:
            days_left = max(
                0,
                (deadline.date() - now.date()).days,
            )

        risks = ""

        if row["risks"]:
            try:
                parsed_risks = json.loads(row["risks"])

                if isinstance(parsed_risks, list):
                    risks = "; ".join(
                        str(x).strip()
                        for x in parsed_risks
                        if str(x).strip()
                    )
                else:
                    risks = str(parsed_risks)

            except (TypeError, ValueError, json.JSONDecodeError):
                risks = str(row["risks"])

        ws.append(
            [
                platform,
                row["external_id"] or "",
                row["title"] or "",
                row["customer"] or "",
                row["region"] or "",
                row["price"],
                row["currency"] or "RUB",
                _excel_datetime(published_at),
                _excel_datetime(deadline),
                days_left,
                row["law_type"] or "",
                procurement_method,
                row["relevance_score"],
                recommendation,
                row["summary"] or "",
                risks,
                row["price"],
                row["deadline_note"] or "",
                row["url"] or "",
            ]
        )

    # Делаем ссылку на тендер настоящей гиперссылкой Excel.
    for cell in ws["S"][1:]:
        if cell.row > 1 and cell.value:
            cell.hyperlink = str(cell.value)
            cell.font = Font(color="0563C1", underline="single")

    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions

    ws.freeze_panes = "A2"

    widths = {
        1: 14,  # Площадка
        2: 22,  # Номер закупки
        3: 55,  # Наименование
        4: 32,  # Заказчик
        5: 18,  # Регион
        6: 18,  # НМЦК
        7: 8,   # Валюта
        8: 14,  # Дата публикации
        9: 24,  # Дата окончания подачи заявок
        10: 20, # Осталось дней до подачи
        11: 8,  # Закон
        12: 30, # Способ закупки
        13: 10, # AI score
        14: 18, # Рекомендация
        15: 45, # Краткое резюме
        16: 35, # Риски
        17: 20, # Начальная цена
        18: 35, # Комментарий по срокам
        19: 55, # Ссылка
    }

    for column, width in widths.items():
        ws.column_dimensions[
            get_column_letter(column)
        ].width = width

    for row_cells in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
    ):
        for cell in row_cells:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    # Форматы Excel:
    # F  = НМЦК
    # H  = дата публикации
    # I  = дата окончания подачи заявок
    # J  = осталось дней до подачи
    # M  = AI score
    # Q  = Начальная цена

    for column in ("H", "I"):
        for cell in ws[column][1:]:
            cell.number_format = "dd.mm.yyyy"

    for column in ("F", "Q"):
        for cell in ws[column][1:]:
            cell.number_format = '#,##0.00 "₽"'

    for cell in ws["J"][1:]:
        cell.number_format = "0"

    for cell in ws["M"][1:]:
        cell.number_format = "0"

    stats = wb.create_sheet("Статистика")

    stats.append(["Показатель", "Значение"])
    stats["A1"].font = Font(bold=True)
    stats["B1"].font = Font(bold=True)

    total = len(rows)

    analyzed = sum(
        1
        for row in rows
        if row["relevance_score"] is not None
    )

    high_score = sum(
        1
        for row in rows
        if row["relevance_score"] is not None
        and int(row["relevance_score"]) >= 70
    )

    participate = sum(
        1
        for row in rows
        if row["recommendation"] == "participate"
    )

    review = sum(
        1
        for row in rows
        if row["recommendation"] == "review"
    )

    skip = sum(
        1
        for row in rows
        if row["recommendation"] == "skip"
    )

    stats_rows = [
        ("Тендеров в этом прогоне", total),
        ("Проанализировано AI", analyzed),
        ("AI score >= 70", high_score),
        ("Рекомендация: участвовать", participate),
        ("Рекомендация: рассмотреть", review),
        ("Рекомендация: пропустить", skip),
        (
            "Дата формирования",
            datetime.now().strftime("%d.%m.%Y %H:%M"),
        ),
    ]

    for item in stats_rows:
        stats.append(item)

    stats.column_dimensions["A"].width = 35
    stats.column_dimensions["B"].width = 25

    for cell in stats["A"][1:]:
        cell.font = Font(bold=True)

    wb.save(output_path)

    return output_path


def _parse_datetime(value) -> datetime | None:
    if not value:
        return None

    if isinstance(value, datetime):
        dt = value
    else:
        try:
            dt = datetime.fromisoformat(str(value))
        except (TypeError, ValueError):
            return None

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    return dt


def _excel_datetime(value: datetime | None):
    if value is None:
        return None

    return value.replace(tzinfo=None)
